import openpyxl as xl, win32com.client, PIL, os, sys, random, datetime, matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference
from PIL import ImageGrab, Image
from docx.shared import Cm
from docx import Document
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu

collage_table_name=['各學院論文數統計', '各學院論文成長率', '各學院被引次數統計', '各系所論文數統計', '各系所被引次數統計']

def table_one(File_name,sheet_name):

    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表1.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': ws.cell(i,2).value,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表一.docx')


def table_two(File_name,sheet_name):

    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表2.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name':  ws.cell(i,2).value ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表二.docx')


def table_three(File_name,sheet_name):
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表3.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name':  ws.cell(i,2).value ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表三.docx')

def table_four(File_name,sheet_name):
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表4.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name':  ws.cell(i,2).value ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表四.docx')


def table_five(File_name,sheet_name): 

    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表5.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': ws.cell(i,2).value ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
          'D2': ws.cell(2,4).value,
          'E2': ws.cell(2,5).value,
          'F2': ws.cell(2,6).value,
          'G2': ws.cell(2,7).value,
          'H2': ws.cell(2,8).value,
          'I2': ws.cell(2,9).value,
          'table_contents': table_contents,
          'Year':datetime.datetime.now().year,
          'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表五.docx')