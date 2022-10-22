import openpyxl as xl, win32com.client, PIL, os, sys, random, datetime, matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference
from PIL import ImageGrab, Image
from docx.shared import Cm
from docx import Document
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu

collage_table_name=['歷年上半年論文數統計','論文數統計','被引次數統計','論文成長率','三年成長率','平均被引用次數統計','H指數']

collage_list_all=['文學院','理學院','社會科學院','管理學院','法律學院','工學院','電機資訊學院','公共衛生學院','生命科學院','生物資源暨農學院','獸醫專業學院','醫學院','牙醫專業學院','藥學專業學院','國家理論科學研究中心(北區)',
'凝態科學研究中心','其他單位','N/A']
collage_list=['文學院','理學院','社會科學院','管理學院','法律學院','工學院','電機資訊學院','公共衛生學院','生命科學院','生物資源暨農學院','獸醫專業學院','醫學院','牙醫專業學院','藥學專業學院']

def table_one(File_name,sheet_name):



    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表1.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表一.docx')


def table_two(File_name,sheet_name):

    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表2.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表二.docx')


def table_three(File_name,sheet_name):
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表3.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表三.docx')

def table_four(File_name,sheet_name):
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表4.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表四.docx')


def table_five(File_name,sheet_name): # 三年成長率
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表5.docx')
    
    table_contents = []

    for i in range(2, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-2] ,
                     'C1': ws.cell(i, 3).value,
                     'D1': ws.cell(i, 4).value,
                     'E1': ws.cell(i, 5).value,
                     'F1': ws.cell(i, 6).value,
                     }
        table_contents.append(table_content)


    context = {
          'C1': ws.cell(1, 3).value,
          'D1': ws.cell(1, 4).value,
          'E1': ws.cell(1, 5).value,
          'F1': ws.cell(1, 6).value,
          'table_contents': table_contents,
          'Year':datetime.datetime.now().year,
          'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表五.docx')


def table_six(File_name,sheet_name):
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表6.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表六.docx')


def table_seven(File_name,sheet_name): #H指數
    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表7.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name': collage_list_all[i-3] ,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save('表七.docx')

