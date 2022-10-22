import openpyxl as xl, win32com.client, PIL, os, sys, random, datetime, matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference
from PIL import ImageGrab, Image
from docx.shared import Cm
from docx import Document
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu

table_name=['文學院', '理學院', '社會科學院', '管理學院', '法律學院', '工學院', '電機資訊學院', '公共衛生學院', '生命科學院', '生物資源暨農學院', '獸醫專業學院',
 '醫學院', '牙醫專業學院', '藥學專業學院', '研究中心', '行政單位']

def table_one(File_name,sheet_name,output_file):



    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表4.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name':ws.cell(i,2).value,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save(output_file+'.docx')


def table_two(File_name,sheet_name,output_file):



    wb = xl.load_workbook(File_name)
    ws = wb[sheet_name]

    template = DocxTemplate('template\表4.docx')
    
    table_contents = []

    for i in range(3, ws.max_row + 1):
        table_content = {
                     'name':ws.cell(i,2).value,
                     'D2': ws.cell(i, 4).value,
                     'E2': ws.cell(i, 5).value,
                     'F2': ws.cell(i, 6).value,
                     'G2': ws.cell(i, 7).value,
                     'H2': ws.cell(i, 8).value,
                     'I2': ws.cell(i, 9).value,
                     'total':ws.cell(i, 10).value,
                     }
        table_contents.append(table_content)


    context = {
         'D2': ws.cell(2,4).value,
         'E2': ws.cell(2,5).value,
         'F2': ws.cell(2,6).value,
         'G2': ws.cell(2,7).value,
         'H2': ws.cell(2,8).value,
         'I2': ws.cell(2,9).value,
         'Depaet':output_file,
         'table_contents': table_contents,
         'Year':datetime.datetime.now().year,
         'between':datetime.datetime.now().year-5,
           }

    template.render(context)
    template.save(output_file+'.docx')